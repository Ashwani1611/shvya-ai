from io import BytesIO
from unittest.mock import patch
from uuid import uuid4

from django.contrib.sessions.backends.db import SessionStore
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.accounts.session_utils import get_session_cookie_name, set_authenticated_user
from apps.channels.models import WhatsAppAccount
from apps.crm.models import AttributeDefinition, Lead, Pipeline, PipelinePermission
from apps.followups.models import FollowupSequence, FollowupStep, LeadSequenceState
from apps.organizations.models import Organization
from services.followup_service import FollowupError, assign_sequence


@override_settings(CACHES={"default": {"BACKEND": "django.core.cache.backends.locmem.LocMemCache"}})
class BulkLeadTests(TestCase):
    def setUp(self):
        self.organization = Organization.objects.create(name="Bulk test")
        self.user = User.objects.create_user(
            email="bulk@example.com", password="test-password", name="Admin",
            organization=self.organization, role=User.Role.ADMIN,
        )
        self.pipeline = self.organization.pipelines.get(name="Leads")
        self.stage = self.pipeline.stages.order_by("display_order").first()
        self.next_stage = self.pipeline.stages.order_by("display_order")[1]
        self.leads = [Lead.objects.create(
            organization=self.organization, pipeline=self.pipeline, stage=self.stage,
            name=f"Lead {i}", phone=f"+9198765432{i:02}", attributes={"company": f"Company {i}"},
        ) for i in range(3)]
        self.authenticate(self.user)

    def authenticate(self, user):
        session = SessionStore()
        set_authenticated_user(session, user)
        session.save()
        self.client.cookies[get_session_cookie_name("dashboard")] = session.session_key

    def post(self, action, **data):
        payload = {
            "action": action, "pipeline": str(self.pipeline.pk), "source_stage": str(self.stage.pk),
            "lead_ids": [str(lead.pk) for lead in self.leads[:2]],
        }
        payload.update(data)
        return self.client.post(reverse("crm-leads-bulk"), payload, content_type="application/json")

    def sequence(self):
        account = WhatsAppAccount.objects.create(
            organization=self.organization, status="connected", is_active=True,
            phone_number_id="12345", display_phone_number="+919999999999",
        )
        sequence = FollowupSequence.objects.create(
            organization=self.organization, name="Nurture", whatsapp_account=account,
        )
        FollowupStep.objects.create(
            sequence=sequence, position=1, step_type="reminder", reminder_text="Call lead",
            schedule_type="delay", delay_value=1, delay_unit="days",
        )
        return sequence

    def test_options_and_rendered_controls_use_live_context_builder(self):
        response = self.post("options")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 2)
        self.assertTrue(response.json()["permissions"]["delete"])
        response = self.client.get(reverse("crm-lead-table-partial"), {
            "pipeline": self.pipeline.pk, "stage": self.stage.pk, "search": "Lead 0",
        })
        self.assertContains(response, 'data-bulk-action="update"')
        self.assertContains(response, 'data-bulk-action="export"')
        self.assertContains(response, 'data-bulk-action="delete"')
        self.assertContains(response, "Select all 1 leads in this stage")
        self.assertContains(response, "search=Lead+0")
        self.assertContains(response, "data-lead-select", count=1)

    def test_moves_only_selected_leads_and_records_activity(self):
        response = self.post("update", move=True, target_pipeline=str(self.pipeline.pk), target_stage=str(self.next_stage.pk))
        self.assertEqual(response.status_code, 200, response.content)
        for lead in self.leads[:2]:
            old_time = lead.stage_entered_at
            lead.refresh_from_db()
            self.assertEqual(lead.stage, self.next_stage)
            self.assertGreater(lead.stage_entered_at, old_time)
            self.assertTrue(lead.activities.filter(topic="stage_changed", actor=self.user).exists())
        self.leads[2].refresh_from_db()
        self.assertEqual(self.leads[2].stage, self.stage)

    def test_same_stage_is_noop(self):
        old_time = self.leads[0].stage_entered_at
        response = self.post("update", move=True, target_pipeline=str(self.pipeline.pk), target_stage=str(self.stage.pk))
        self.assertEqual(response.status_code, 200)
        self.leads[0].refresh_from_db()
        self.assertEqual(self.leads[0].stage_entered_at, old_time)
        self.assertFalse(self.leads[0].activities.exists())

    def test_pipeline_move_preserves_attributes_and_history(self):
        target = Pipeline.objects.create(organization=self.organization, name="Sales")
        stage = target.stages.first()
        response = self.post("update", move=True, target_pipeline=str(target.pk), target_stage=str(stage.pk))
        self.assertEqual(response.status_code, 200, response.content)
        self.leads[0].refresh_from_db()
        self.assertEqual(self.leads[0].pipeline, target)
        self.assertEqual(self.leads[0].attributes, {"company": "Company 0"})
        self.assertTrue(self.leads[0].activities.filter(topic="pipeline_changed").exists())

    def test_wrong_destination_stage_and_inactive_stage_rejected(self):
        target = Pipeline.objects.create(organization=self.organization, name="Sales")
        response = self.post("update", move=True, target_pipeline=str(target.pk), target_stage=str(self.stage.pk))
        self.assertEqual(response.status_code, 400)
        self.next_stage.is_active = False
        self.next_stage.save()
        response = self.post("update", move=True, target_pipeline=str(self.pipeline.pk), target_stage=str(self.next_stage.pk))
        self.assertEqual(response.status_code, 400)

    def test_other_organization_is_never_exported_or_modified(self):
        organization = Organization.objects.create(name="Other org")
        pipeline = organization.pipelines.first()
        lead = Lead.objects.create(organization=organization, pipeline=pipeline, stage=pipeline.stages.first(), name="Private", phone="+919999999998")
        for action in ("options", "export", "update", "delete"):
            response = self.post(action, lead_ids=[str(self.leads[0].pk), str(lead.pk)], confirm_delete=True)
            self.assertEqual(response.status_code, 400)
        self.assertTrue(Lead.objects.filter(pk=lead.pk).exists())
        self.assertEqual(Lead.objects.filter(organization=self.organization).count(), 3)

    def test_empty_malformed_and_stale_selections_rejected(self):
        for ids in ([], ["invalid"], [str(uuid4())], "not-a-list"):
            self.assertEqual(self.post("export", lead_ids=ids).status_code, 400)
        self.leads[0].stage = self.next_stage
        self.leads[0].save()
        self.assertEqual(self.post("delete", confirm_delete=True).status_code, 400)
        self.assertEqual(Lead.objects.filter(organization=self.organization).count(), 3)

    def test_duplicate_ids_are_deduplicated(self):
        response = self.post("options", lead_ids=[str(self.leads[0].pk)] * 2)
        self.assertEqual(response.json()["count"], 1)

    def test_agent_permissions_are_checked_for_each_action(self):
        agent = User.objects.create_user(email="agent@example.com", name="Agent", password="test", organization=self.organization, role=User.Role.AGENT)
        self.pipeline.owner = agent
        self.pipeline.save()
        self.authenticate(agent)
        self.assertEqual(self.post("export").status_code, 200)
        self.assertEqual(self.post("delete", confirm_delete=True).status_code, 403)
        self.assertEqual(self.post("update", move=True).status_code, 403)
        self.assertEqual(self.post("update", sequence_action="clear").status_code, 403)
        permission = PipelinePermission.objects.create(user=agent, pipeline=self.pipeline, can_delete_leads=True)
        self.assertTrue(permission.can_delete_leads)
        self.assertEqual(self.post("delete", confirm_delete=True).status_code, 200)

    def test_agent_cannot_bypass_pipeline_access_using_filters(self):
        agent = User.objects.create_user(email="agent@example.com", name="Agent", password="test", organization=self.organization, role=User.Role.AGENT)
        self.pipeline.owner = agent
        self.pipeline.save()
        target = Pipeline.objects.create(organization=self.organization, name="Private pipeline")
        self.authenticate(agent)
        response = self.client.get(reverse("crm-lead-table-partial"), {"pipeline": self.pipeline.pk, "filter_pipeline": target.pk})
        self.assertEqual(response.context["selected_pipeline_id"], str(self.pipeline.pk))
        self.assertNotContains(response, 'data-bulk-action="delete"')
        self.assertEqual(self.post("export", pipeline=str(target.pk)).status_code, 403)

    def test_delete_requires_confirmation_and_removes_only_selection(self):
        self.assertEqual(self.post("delete").status_code, 400)
        response = self.post("delete", confirm_delete=True)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 2)
        self.assertEqual(list(Lead.objects.filter(organization=self.organization)), [self.leads[2]])

    def test_noop_update_rejected(self):
        self.assertEqual(self.post("update").status_code, 400)

    def test_sequence_assignment_and_clearing_use_existing_scheduler(self):
        sequence = self.sequence()
        response = self.post("update", sequence_action="assign", sequence=str(sequence.pk))
        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(LeadSequenceState.objects.filter(status="active", next_step__isnull=False).count(), 2)
        response = self.post("update", sequence_action="clear")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(LeadSequenceState.objects.filter(status="cleared", next_step__isnull=True, upcoming_send_at__isnull=True).count(), 2)

    def test_later_sequence_failure_rolls_back_moves_and_assignments(self):
        sequence = self.sequence()
        attempts = []

        def assign_then_fail(**kwargs):
            attempts.append(kwargs["lead"].pk)
            if len(attempts) == 2:
                raise FollowupError("Sender mismatch")
            return assign_sequence(**kwargs)

        with patch("apps.crm.views.bulk.assign_sequence", side_effect=assign_then_fail):
            response = self.post("update", move=True, target_pipeline=str(self.pipeline.pk), target_stage=str(self.next_stage.pk), sequence_action="assign", sequence=str(sequence.pk))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(len(attempts), 2)
        self.assertEqual(Lead.objects.filter(stage=self.stage).count(), 3)
        self.assertFalse(LeadSequenceState.objects.exists())
        self.assertFalse(self.leads[0].activities.exists())

    def test_sender_mismatch_rolls_back_pipeline_move(self):
        sequence = self.sequence()
        target = Pipeline.objects.create(organization=self.organization, name="Different sender", phone_number="+918888888888")
        response = self.post("update", move=True, target_pipeline=str(target.pk), target_stage=str(target.stages.first().pk), sequence_action="assign", sequence=str(sequence.pk))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Lead.objects.filter(pipeline=self.pipeline).count(), 3)

    def test_export_all_attributes_includes_custom_and_literal_phone(self):
        AttributeDefinition.objects.create(organization=self.organization, name="Budget", key="budget")
        response = self.post("export")
        self.assertEqual(response.status_code, 200)
        self.assertIn(".xlsx", response["Content-Disposition"])
        sheet = load_workbook(BytesIO(response.content)).active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Budget (custom)", headers)
        self.assertIn("company (custom)", headers)
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual({row[headers.index("Phone")] for row in sheet.iter_rows(min_row=2, values_only=True)}, {lead.phone for lead in self.leads[:2]})

    def test_export_selected_fields_and_formula_safety(self):
        self.leads[0].name = '=HYPERLINK("https://example.com")'
        self.leads[0].save()
        response = self.post("export", attribute_mode="selected", attributes=["name", "attr:company"])
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.content)).active
        self.assertEqual([cell.value for cell in sheet[1]], ["Name", "company (custom)"])
        self.assertEqual(sheet.max_column, 2)
        cell = next(row[0] for row in sheet.iter_rows(min_row=2) if row[0].value.startswith("="))
        self.assertEqual(cell.data_type, "s")
        for attributes in ([], ["unknown"], [{}]):
            self.assertEqual(self.post("export", attribute_mode="selected", attributes=attributes).status_code, 400)

    def test_authentication_and_post_required(self):
        self.assertEqual(self.client.get(reverse("crm-leads-bulk")).status_code, 405)
        self.client.cookies.clear()
        self.assertEqual(self.post("export").status_code, 302)
