"""Organization-scoped bulk CRM actions; all mutations commit together."""

import json
from io import BytesIO
from uuid import UUID

from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

from apps.accounts.models import User
from apps.crm.decorators import crm_login_required
from apps.crm.models import AttributeDefinition, Lead, PipelinePermission, Stage
from apps.followups.models import FollowupSequence
from services.crm.lead_transition import LeadTransitionError, move_lead_to_stage
from services.crm_activity_service import record_pipeline_changed
from services.followup_service import FollowupError, assign_sequence, clear_sequence

from .api import get_user_pipelines


def bulk_permissions(user, pipeline):
    if user.role == User.Role.ADMIN:
        return {"move": True, "edit": True, "delete": True}
    permission = PipelinePermission.objects.filter(user=user, pipeline=pipeline).first()
    return {
        action: bool(permission and getattr(permission, field))
        for action, field in {
            "move": "can_move_leads", "edit": "can_edit_leads", "delete": "can_delete_leads",
        }.items()
    }


def _uuid(value):
    try:
        return UUID(str(value))
    except (ValueError, TypeError, AttributeError) as exc:
        raise ValueError("The selection is invalid. Refresh the CRM and select the leads again.") from exc


def _selection(user, data, *, lock=False):
    ids = data.get("lead_ids")
    if not isinstance(ids, list) or not ids:
        raise ValueError("Select at least one lead.")
    ids = {_uuid(value) for value in ids}
    pipeline = get_user_pipelines(user).filter(pk=_uuid(data.get("pipeline"))).first()
    if pipeline is None:
        raise PermissionDenied
    queryset = Lead.objects.filter(
        pk__in=ids, organization=user.organization, pipeline=pipeline,
        stage_id=_uuid(data.get("source_stage")),
    ).select_related("pipeline", "stage", "organization").order_by("pk")
    if lock:
        queryset = queryset.select_for_update(of=("self",))
    leads = list(queryset)
    if len(leads) != len(ids):
        raise ValueError("Some selected leads have moved, been deleted, or are no longer accessible. Refresh and select them again.")
    return pipeline, leads


CORE_FIELDS = [
    ("id", "Lead ID"), ("name", "Name"), ("phone", "Phone"), ("email", "Email"),
    ("pipeline", "Pipeline"), ("stage", "Stage"), ("notes", "Notes"),
    ("lead_source", "Lead source"), ("ai_enabled", "AI enabled"),
    ("created_at", "Created at (UTC)"), ("updated_at", "Updated at (UTC)"),
    ("stage_entered_at", "Stage entered at (UTC)"),
]


def _fields(user, leads):
    fields = list(CORE_FIELDS)
    definitions = dict(AttributeDefinition.objects.filter(
        organization=user.organization,
    ).values_list("key", "name"))
    for lead in leads:
        for key in (lead.attributes or {}):
            definitions.setdefault(key, key)
    fields.extend((f"attr:{key}", f"{name} (custom)") for key, name in definitions.items())
    return fields


def _options(user, pipeline, leads):
    pipelines = []
    for target in get_user_pipelines(user).prefetch_related("stages"):
        if not bulk_permissions(user, target)["move"]:
            continue
        pipelines.append({
            "id": str(target.pk), "name": target.name,
            "stages": [{"id": str(stage.pk), "name": stage.name}
                       for stage in target.stages.all() if stage.is_active],
        })
    sequences = FollowupSequence.objects.filter(
        organization=user.organization, is_active=True,
        whatsapp_account__organization=user.organization,
        whatsapp_account__is_active=True, whatsapp_account__status="connected",
    ).order_by("name")
    return JsonResponse({
        "count": len(leads), "permissions": bulk_permissions(user, pipeline),
        "pipelines": pipelines,
        "sequences": [{"id": str(s.pk), "name": s.name} for s in sequences],
        "fields": [{"id": key, "name": name} for key, name in _fields(user, leads)],
    })


def _export(user, leads, data):
    fields = _fields(user, leads)
    mode = data.get("attribute_mode", "all")
    if mode == "selected":
        selected = data.get("attributes")
        valid = {key for key, _ in fields}
        if not isinstance(selected, list) or not selected or any(
            not isinstance(key, str) or key not in valid for key in selected
        ):
            raise ValueError("Choose at least one valid attribute to export.")
        fields = [(key, name) for key, name in fields if key in selected]
    elif mode != "all":
        raise ValueError("Choose all attributes or selected attributes.")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append([name for _, name in fields])
    for cell in sheet[1]:
        cell.data_type = "s"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3B82F6")
    for lead in leads:
        row = []
        for key, _ in fields:
            if key.startswith("attr:"):
                value = (lead.attributes or {}).get(key[5:], "")
            elif key in ("pipeline", "stage"):
                value = getattr(lead, key).name
            else:
                value = getattr(lead, key)
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            elif isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            elif value is not None and not isinstance(value, (str, int, float, bool)):
                value = str(value)
            if isinstance(value, str):
                value = ILLEGAL_CHARACTERS_RE.sub("", value)
            row.append(value)
        sheet.append(row)
        # User data is literal text, never an Excel formula (including '=' names).
        for cell in sheet[sheet.max_row]:
            if isinstance(cell.value, str):
                cell.data_type = "s"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            50, max(16, max(len(str(cell.value or "")) for cell in column) + 2),
        )
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="leads-{timezone.now():%Y%m%d-%H%M%S}.xlsx"'
    response["Cache-Control"] = "no-store"
    return response


def _update(user, pipeline, leads, data):
    rights = bulk_permissions(user, pipeline)
    move = data.get("move") is True
    sequence_action = data.get("sequence_action", "keep")
    if sequence_action not in ("keep", "assign", "clear"):
        raise ValueError("Choose a valid sequence action.")
    if not move and sequence_action == "keep":
        raise ValueError("Choose at least one update to apply.")
    if (move and not rights["move"]) or (sequence_action != "keep" and not rights["edit"]):
        raise PermissionDenied
    target, stage, sequence = None, None, None
    if move:
        target = get_user_pipelines(user).filter(pk=_uuid(data.get("target_pipeline"))).first()
        if target is None or not bulk_permissions(user, target)["move"]:
            raise PermissionDenied
        stage = Stage.objects.filter(pk=_uuid(data.get("target_stage")), pipeline=target, is_active=True).first()
        if stage is None:
            raise ValueError("Choose an active stage in the destination pipeline.")
    if sequence_action == "assign":
        sequence = FollowupSequence.objects.filter(
            pk=_uuid(data.get("sequence")), organization=user.organization, is_active=True,
        ).select_related("whatsapp_account").first()
        if sequence is None:
            raise ValueError("Choose an active Auto Followup sequence.")
    for lead in leads:
        if move:
            if lead.pipeline_id == target.pk:
                move_lead_to_stage(lead=lead, stage=stage, actor=user)
            else:
                old_pipeline, old_stage = lead.pipeline, lead.stage
                lead.pipeline, lead.stage = target, stage
                lead.stage_entered_at = timezone.now()
                lead.full_clean()
                lead.save(update_fields=["pipeline", "stage", "stage_entered_at", "updated_at"])
                record_pipeline_changed(lead=lead, actor=user, old_pipeline=old_pipeline,
                                        new_pipeline=target, old_stage=old_stage, new_stage=stage)
        if sequence_action == "assign":
            assign_sequence(lead=lead, sequence=sequence, actor=user)
        elif sequence_action == "clear":
            clear_sequence(lead=lead)


@crm_login_required
@require_POST
def bulk_leads(request):
    try:
        data = json.loads(request.body)
        if not isinstance(data, dict):
            return JsonResponse({"error": "Invalid request."}, status=400)
        action = data.get("action")
        if action not in ("options", "update", "export", "delete"):
            raise ValueError("Choose a valid bulk action.")
        with transaction.atomic():
            pipeline, leads = _selection(request.crm_user, data, lock=action in ("update", "delete"))
            if action == "options":
                return _options(request.crm_user, pipeline, leads)
            if action == "export":
                return _export(request.crm_user, leads, data)
            if action == "update":
                _update(request.crm_user, pipeline, leads, data)
            else:
                if not bulk_permissions(request.crm_user, pipeline)["delete"]:
                    raise PermissionDenied
                if data.get("confirm_delete") is not True:
                    raise ValueError("Confirm permanent deletion of the selected leads.")
                Lead.objects.filter(pk__in=[lead.pk for lead in leads]).delete()
        return JsonResponse({"count": len(leads), "action": action})
    except PermissionDenied:
        return JsonResponse({"error": "You do not have permission to perform this action on these leads."}, status=403)
    except ProtectedError:
        return JsonResponse({"error": "These leads have protected records and cannot be deleted. No leads were deleted."}, status=409)
    except (ValueError, ValidationError, LeadTransitionError, FollowupError) as exc:
        message = " ".join(exc.messages) if isinstance(exc, ValidationError) else str(exc)
        return JsonResponse({"error": message}, status=400)
