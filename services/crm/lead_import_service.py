"""
Lead import parsing and temporary wizard state.

This service is responsible only for:

    - reading uploaded CSV/XLS/XLSX files
    - validating import files
    - normalizing headers/cells
    - normalizing imported phone numbers
    - storing temporary Import Leads wizard state in Redis

Lead creation/update logic remains in:
    services.crm.lead_service

Attribute definition logic remains in:
    services.crm.attribute_service
"""

import csv
import io
import re
import uuid
from pathlib import Path

from django.core.cache import cache
from django.core.exceptions import ValidationError

from openpyxl import load_workbook
import xlrd


# ============================================================
# IMPORT CONSTANTS
# ============================================================

MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

MAX_IMPORT_ROWS = 25_000

IMPORT_STATE_TIMEOUT = 60 * 60  # 1 hour

IMPORT_CACHE_PREFIX = "shvya:lead_import:"

SUPPORTED_EXTENSIONS = {
    ".csv",
    ".xls",
    ".xlsx",
}


# ============================================================
# NORMALIZATION
# ============================================================


def _normalize_header(
    value,
):
    """
    Normalize a spreadsheet header.

    Example:

        " First Name " -> "First Name"
    """

    if value is None:

        return ""

    return str(
        value
    ).strip()


def _normalize_cell(
    value,
):
    """
    Convert spreadsheet cell values into strings suitable
    for the import wizard.

    Empty values become an empty string.
    """

    if value is None:

        return ""

    if isinstance(
        value,
        float,
    ):

        if value.is_integer():

            return str(
                int(value)
            )

    return str(
        value
    ).strip()


# ============================================================
# IMPORT PHONE NORMALIZATION
# ============================================================


def normalize_import_phone(
    value,
):
    """
    Normalize phone numbers coming from imported
    CSV/XLS/XLSX files.

    Supported examples:

        +919164543255
        +91 9164543255
        +91-9164543255
        919164543255
        p:+919164543255
        p:9164543255
        p:+91**9164543255

    All of the above normalize to:

        +919164543255

    Other international country codes are also supported,
    for example:

        +1
        +44
        +61
        +81
        +971

    This function performs import formatting normalization
    only.

    It does not perform:

        - carrier validation
        - live-number verification
        - country-specific subscriber validation

    The importer treats the international country code as
    already being present when supplied in the source value.

    Special rule:

        p:9164543255

    is interpreted as an Indian mobile number because the
    explicit `p:` prefix is present and the remaining value
    is a 10-digit Indian mobile-style number.
    """

    original_value = _normalize_cell(
        value
    )

    if not original_value:

        return ""

    # --------------------------------------------------------
    # DETECT THE SPECIAL p: PREFIX
    # --------------------------------------------------------

    had_p_prefix = (
        original_value
        .lower()
        .startswith("p:")
    )

    value = original_value

    if had_p_prefix:

        value = value[2:].strip()

    # --------------------------------------------------------
    # KEEP DIGITS ONLY
    #
    # Examples:
    #
    #     +91 9164543255
    #         -> 919164543255
    #
    #     +91-9164543255
    #         -> 919164543255
    #
    #     +91**9164543255
    #         -> 919164543255
    #
    #     p:+91**9164543255
    #         -> 919164543255
    # --------------------------------------------------------

    digits = re.sub(
        r"\D",
        "",
        value,
    )

    if not digits:

        raise ValidationError(
            {
                "phone": (
                    "Phone number must contain digits."
                )
            }
        )

    # --------------------------------------------------------
    # SPECIAL p: INDIAN MOBILE RULE
    #
    # Only an explicit p: prefix triggers this rule.
    #
    #     p:9164543255
    #
    # becomes:
    #
    #     +919164543255
    #
    # This prevents us from assuming that every 10-digit
    # number belongs to India.
    # --------------------------------------------------------

    if (
        had_p_prefix
        and len(digits) == 10
        and digits[0] in "6789"
    ):

        digits = (
            "91"
            + digits
        )

    # --------------------------------------------------------
    # BASIC LENGTH CHECK
    # --------------------------------------------------------

    if len(digits) < 8:

        raise ValidationError(
            {
                "phone": (
                    "Phone number is too short."
                )
            }
        )

    # --------------------------------------------------------
    # E.164 MAXIMUM DIGIT LENGTH
    #
    # '+' is not counted.
    # --------------------------------------------------------

    if len(digits) > 15:

        raise ValidationError(
            {
                "phone": (
                    "Phone number cannot contain more "
                    "than 15 digits."
                )
            }
        )

    # --------------------------------------------------------
    # VALID COUNTRY-CODE STYLE PREFIX
    #
    # We do not maintain a hard-coded list of country codes
    # here. The importer only requires that the supplied
    # international digits have a plausible country-code
    # prefix length.
    #
    # Supported examples include:
    #
    #     +1
    #     +44
    #     +61
    #     +81
    #     +91
    #     +971
    #
    # A 1-to-3 digit international prefix is expected.
    # --------------------------------------------------------

    if len(digits) < 9:

        raise ValidationError(
            {
                "phone": (
                    "Phone number must contain a country "
                    "code and subscriber number."
                )
            }
        )

    # --------------------------------------------------------
    # RETURN CANONICAL SHVYA FORMAT
    # --------------------------------------------------------

    return (
        "+"
        + digits
    )


# ============================================================
# FILE VALIDATION
# ============================================================


def _validate_uploaded_file(
    uploaded_file,
):
    """
    Validate uploaded file size and extension.

    Returns:
        normalized file extension
    """

    if uploaded_file is None:

        raise ValidationError(
            {
                "file": (
                    "Please select a file to upload."
                )
            }
        )

    if uploaded_file.size <= 0:

        raise ValidationError(
            {
                "file": (
                    "The uploaded file is empty."
                )
            }
        )

    if uploaded_file.size > MAX_IMPORT_FILE_SIZE:

        raise ValidationError(
            {
                "file": (
                    "File size cannot exceed 10 MB."
                )
            }
        )

    filename = (
        getattr(
            uploaded_file,
            "name",
            "",
        )
        or ""
    )

    extension = (
        Path(
            filename
        )
        .suffix
        .lower()
    )

    if extension not in SUPPORTED_EXTENSIONS:

        raise ValidationError(
            {
                "file": (
                    "Unsupported file type. "
                    "Please upload CSV, XLS, or XLSX."
                )
            }
        )

    return extension


# ============================================================
# HEADER HANDLING
# ============================================================


def _ensure_unique_headers(
    headers,
):
    """
    Make duplicate spreadsheet headers unique.

    Example:

        Email
        Email

    becomes:

        Email
        Email (2)
    """

    seen = {}

    normalized = []

    for header in headers:

        base = _normalize_header(
            header
        )

        if not base:

            base = "Unnamed Column"

        count = (
            seen.get(
                base,
                0,
            )
            + 1
        )

        seen[base] = count

        if count == 1:

            normalized.append(
                base
            )

        else:

            normalized.append(
                f"{base} ({count})"
            )

    return normalized


# ============================================================
# CSV
# ============================================================


def _parse_csv(
    uploaded_file,
):
    """
    Parse CSV content.

    Returns:
        headers, rows
    """

    uploaded_file.seek(0)

    raw = uploaded_file.read()

    if isinstance(
        raw,
        bytes,
    ):

        text = raw.decode(
            "utf-8-sig",
            errors="replace",
        )

    else:

        text = str(
            raw
        )

    reader = csv.reader(
        io.StringIO(
            text
        )
    )

    try:

        raw_headers = next(
            reader
        )

    except StopIteration:

        raise ValidationError(
            {
                "file": (
                    "The CSV file contains no rows."
                )
            }
        )

    headers = _ensure_unique_headers(
        raw_headers
    )

    rows = []

    for row in reader:

        values = [
            _normalize_cell(
                value
            )
            for value in row
        ]

        if len(values) < len(headers):

            values.extend(
                [""] * (
                    len(headers)
                    - len(values)
                )
            )

        elif len(values) > len(headers):

            values = values[
                :len(headers)
            ]

        if not any(values):

            continue

        rows.append(
            dict(
                zip(
                    headers,
                    values,
                )
            )
        )

    return (
        headers,
        rows,
    )


# ============================================================
# XLSX
# ============================================================


def _parse_xlsx(
    uploaded_file,
):
    """
    Parse XLSX workbook using openpyxl.

    The first worksheet is used.
    """

    uploaded_file.seek(0)

    workbook = load_workbook(
        filename=uploaded_file,
        read_only=True,
        data_only=True,
    )

    try:

        if not workbook.worksheets:

            raise ValidationError(
                {
                    "file": (
                        "The XLSX file contains "
                        "no worksheets."
                    )
                }
            )

        worksheet = (
            workbook.worksheets[0]
        )

        iterator = worksheet.iter_rows(
            values_only=True
        )

        try:

            raw_headers = next(
                iterator
            )

        except StopIteration:

            raise ValidationError(
                {
                    "file": (
                        "The XLSX file contains "
                        "no rows."
                    )
                }
            )

        headers = _ensure_unique_headers(
            raw_headers
        )

        rows = []

        for raw_row in iterator:

            values = [
                _normalize_cell(
                    value
                )
                for value in raw_row
            ]

            if len(values) < len(headers):

                values.extend(
                    [""] * (
                        len(headers)
                        - len(values)
                    )
                )

            elif len(values) > len(headers):

                values = values[
                    :len(headers)
                ]

            if not any(values):

                continue

            rows.append(
                dict(
                    zip(
                        headers,
                        values,
                    )
                )
            )

        return (
            headers,
            rows,
        )

    finally:

        workbook.close()


# ============================================================
# XLS
# ============================================================


def _parse_xls(
    uploaded_file,
):
    """
    Parse legacy XLS workbook using xlrd.

    The first worksheet is used.
    """

    uploaded_file.seek(0)

    workbook = xlrd.open_workbook(
        file_contents=uploaded_file.read(),
        on_demand=True,
    )

    try:

        if workbook.nsheets == 0:

            raise ValidationError(
                {
                    "file": (
                        "The XLS file contains "
                        "no worksheets."
                    )
                }
            )

        worksheet = (
            workbook.sheet_by_index(0)
        )

        if worksheet.nrows == 0:

            raise ValidationError(
                {
                    "file": (
                        "The XLS file contains "
                        "no rows."
                    )
                }
            )

        raw_headers = (
            worksheet.row_values(
                0
            )
        )

        headers = _ensure_unique_headers(
            raw_headers
        )

        rows = []

        for row_index in range(
            1,
            worksheet.nrows,
        ):

            raw_row = (
                worksheet.row_values(
                    row_index
                )
            )

            values = [
                _normalize_cell(
                    value
                )
                for value in raw_row
            ]

            if len(values) < len(headers):

                values.extend(
                    [""] * (
                        len(headers)
                        - len(values)
                    )
                )

            elif len(values) > len(headers):

                values = values[
                    :len(headers)
                ]

            if not any(values):

                continue

            rows.append(
                dict(
                    zip(
                        headers,
                        values,
                    )
                )
            )

        return (
            headers,
            rows,
        )

    finally:

        workbook.release_resources()


# ============================================================
# MAIN PARSER
# ============================================================


def parse_uploaded_file(
    uploaded_file,
):
    """
    Validate and parse a CSV, XLS, or XLSX file.

    Returns:

        {
            "filename": str,
            "extension": str,
            "headers": list[str],
            "rows": list[dict[str, str]],
            "row_count": int,
        }
    """

    extension = (
        _validate_uploaded_file(
            uploaded_file
        )
    )

    if extension == ".csv":

        headers, rows = (
            _parse_csv(
                uploaded_file
            )
        )

    elif extension == ".xlsx":

        headers, rows = (
            _parse_xlsx(
                uploaded_file
            )
        )

    elif extension == ".xls":

        headers, rows = (
            _parse_xls(
                uploaded_file
            )
        )

    else:

        raise ValidationError(
            {
                "file": (
                    "Unsupported import file."
                )
            }
        )

    if not headers:

        raise ValidationError(
            {
                "file": (
                    "The uploaded file must contain "
                    "at least one column."
                )
            }
        )

    if len(rows) > MAX_IMPORT_ROWS:

        raise ValidationError(
            {
                "file": (
                    f"The file contains more than "
                    f"{MAX_IMPORT_ROWS:,} data rows. "
                    "Please split the file into smaller files."
                )
            }
        )

    return {
        "filename": (
            getattr(
                uploaded_file,
                "name",
                "",
            )
            or ""
        ),
        "extension": extension,
        "headers": headers,
        "rows": rows,
        "row_count": len(rows),
    }


# ============================================================
# TEMPORARY IMPORT STATE
# ============================================================


def create_import_token():
    """
    Create a random token identifying one Import Leads
    wizard session.
    """

    return uuid.uuid4().hex


def _import_cache_key(
    import_token,
):
    """
    Build the Redis cache key for one import session.
    """

    return (
        f"{IMPORT_CACHE_PREFIX}{import_token}"
    )


def save_import_state(
    import_token,
    state,
):
    """
    Save temporary Import Leads wizard state in Redis.

    State automatically expires after IMPORT_STATE_TIMEOUT.
    """

    cache.set(
        _import_cache_key(
            import_token
        ),
        state,
        timeout=IMPORT_STATE_TIMEOUT,
    )


def get_import_state(
    import_token,
):
    """
    Retrieve temporary Import Leads wizard state.

    Returns None when:

        - token is empty
        - token does not exist
        - state has expired
    """

    if not import_token:

        return None

    return cache.get(
        _import_cache_key(
            import_token
        )
    )


def delete_import_state(
    import_token,
):
    """
    Delete temporary Import Leads wizard state.
    """

    if not import_token:

        return

    cache.delete(
        _import_cache_key(
            import_token
        )
    )